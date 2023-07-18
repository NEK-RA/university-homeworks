import openpyxl

class Book:
  def __init__(self, path):
    self.xls = openpyxl.load_workbook(path)
    self.sheet = self.xls.active
    self.cols = self.sheet.max_column
    self.rows = self.sheet.max_row
  
  def cell(self, row, column):
    if row-1 in range(self.rows) and column-1 in range(self.cols):
      return self.sheet.cell(row=row, column=column).value
    else:
      return f"Некорректная ячейка: индексы должны начинаться с 1 и заканчиваться {self.rows} для строк или {self.cols} для столбцов"